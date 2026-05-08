import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from datetime import date
import streamlit.components.v1 as components

st.set_page_config(page_title="CMDB Pregled", layout="wide")
st.title("📊 CMDB Pregled")

# =========================
# SESSION STATE
# =========================
if "transfer_list" not in st.session_state:
    st.session_state.transfer_list = []

if "generated_excel" not in st.session_state:
    st.session_state.generated_excel = None

if "generated_file_name" not in st.session_state:
    st.session_state.generated_file_name = ""

if "print_html" not in st.session_state:
    st.session_state.print_html = ""

if "last_chance_results" not in st.session_state:
    st.session_state.last_chance_results = pd.DataFrame()

# =========================
# LOAD DATA
# =========================
@st.cache_data
def load_data():
    try:
        return pd.read_excel("data.xlsx", dtype=str).fillna("")
    except:
        return pd.DataFrame()

df = load_data()

if df.empty:
    st.warning("data.xlsx nije pronađen ili je prazan")
    st.stop()

# =========================
# HELPERS
# =========================
def set_cell(ws, cell, value):
    for merged_range in ws.merged_cells.ranges:
        if cell in merged_range:
            top_left = merged_range.start_cell.coordinate
            ws[top_left] = value
            ws[top_left].alignment = Alignment(horizontal="center", vertical="center")
            return

    ws[cell] = value
    ws[cell].alignment = Alignment(horizontal="center", vertical="center")


def to_excel(dataframe):
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        dataframe.to_excel(writer, index=False, sheet_name="CMDB")
    return output.getvalue()


def build_print_html(selected_rows, transfer_type):
    if transfer_type == "BG_NS":
        naslov = "Interni prenos BG → NS"
        iz_magacina = "FSBG"
        zaduzio = "FSNS"
    else:
        naslov = "Interni prenos NIŠ → NS"
        iz_magacina = "FSNIŠ"
        zaduzio = "FSNS"

    rows_html = ""

    for i, row in enumerate(selected_rows, start=1):
        rows_html += f"""
        <tr>
            <td>{i}</td>
            <td>{row.get("Name", "")}</td>
            <td>{row.get("Model", "")}</td>
            <td>{row.get("InventoryNumber", "")}</td>
            <td>{row.get("SerialNumber", "")}</td>
            <td>{row.get("SPInventoryNumber", "")}</td>
        </tr>
        """

    return f"""
    <html>
    <head>
        <style>
            body {{ font-family: Arial; padding: 30px; }}
            h2 {{ text-align: center; }}
            table {{ width: 100%; border-collapse: collapse; margin-top: 20px; }}
            td, th {{ border: 1px solid black; padding: 6px; text-align: center; }}
        </style>
    </head>
    <body>
        <button onclick="window.print()">🖨️ Print</button>
        <h2>{naslov}</h2>
        <p><b>Datum:</b> {date.today().strftime("%d.%m.%Y")}</p>
        <p><b>Iz magacina:</b> {iz_magacina}</p>
        <p><b>Uređaj zadužio:</b> {zaduzio}</p>

        <table>
            <tr>
                <th>BR</th>
                <th>NAZIV</th>
                <th>MODEL</th>
                <th>INV</th>
                <th>SN</th>
                <th>SP</th>
            </tr>
            {rows_html}
        </table>
    </body>
    </html>
    """


def generate_internal_transfer(selected_rows, transfer_type):
    if not selected_rows:
        st.error("Lista je prazna.")
        return

    if transfer_type == "BG_NS":
        broj_prenosa = "BG-NS"
        iz_magacina = "FSBG"
        uredjaj_zaduzio = "FSNS"
        file_name = "interni_prenos_BG_NS.xlsx"
        iz_magacina_cell = "B8"
    else:
        broj_prenosa = "FSNIS-FSNS"
        iz_magacina = "FSNIŠ"
        uredjaj_zaduzio = "FSNS"
        file_name = "interni_prenos_NIS_NS.xlsx"
        iz_magacina_cell = "C8"

    try:
        wb = load_workbook("otpremnica_template.xlsx")
        ws = wb.active
    except:
        st.error("Nije pronađen fajl: otpremnica_template.xlsx")
        return

    set_cell(ws, "F4", broj_prenosa)
    set_cell(ws, "G5", date.today().strftime("%d.%m.%Y"))
    set_cell(ws, iz_magacina_cell, iz_magacina)
    set_cell(ws, "G8", uredjaj_zaduzio)

    for i, row in enumerate(selected_rows, start=1):
        r = 14 + i - 1
        set_cell(ws, f"B{r}", i)
        set_cell(ws, f"C{r}", row.get("Name", ""))
        set_cell(ws, f"D{r}", row.get("Model", ""))
        set_cell(ws, f"E{r}", row.get("InventoryNumber", ""))
        set_cell(ws, f"F{r}", row.get("SerialNumber", ""))
        set_cell(ws, f"G{r}", row.get("SPInventoryNumber", ""))

    out = BytesIO()
    wb.save(out)

    st.session_state.generated_excel = out.getvalue()
    st.session_state.generated_file_name = file_name
    st.session_state.print_html = build_print_html(selected_rows, transfer_type)


def add_selected(selected_rows):
    if selected_rows.empty:
        st.warning("Nisi izabrao nijedan uređaj.")
        return

    added = 0
    existing_sp = [
        x.get("SPInventoryNumber", "")
        for x in st.session_state.transfer_list
    ]

    for _, row in selected_rows.iterrows():
        sp = row.get("SPInventoryNumber", "")

        if sp and sp not in existing_sp:
            st.session_state.transfer_list.append(row.to_dict())
            existing_sp.append(sp)
            added += 1

    if added > 0:
        st.success(f"Dodato uređaja: {added}")
    else:
        st.warning("Nema novih uređaja za dodavanje.")

# =========================
# SEARCH
# =========================
st.subheader("🔎 Pretraga")

SEARCH_COLUMNS = [
    "SPInventoryNumber",
    "Name",
    "Vendor",
    "Model",
    "Type",
    "InventoryNumber",
    "SerialNumber"
]

available_search_columns = [c for c in SEARCH_COLUMNS if c in df.columns]

search_col = st.selectbox(
    "Parametar",
    available_search_columns,
    index=available_search_columns.index("SPInventoryNumber") if "SPInventoryNumber" in available_search_columns else 0
)

search_value = st.text_input("Vrednost")

if search_value:
    filtered_df = df[
        df[search_col].astype(str).str.contains(search_value, case=False, na=False)
    ].copy()

    st.subheader(f"📦 Rezultati: {len(filtered_df)}")

    if filtered_df.empty:
        st.info("Nema rezultata.")
    else:
        display_cols = [
            "Name", "Vendor", "Model", "Type",
            "SPInventoryNumber", "InventoryNumber", "SerialNumber"
        ]

        available_cols = [c for c in display_cols if c in filtered_df.columns]
        view_df = filtered_df[available_cols].copy()

        event = st.dataframe(
            view_df,
            use_container_width=True,
            hide_index=True,
            height=300,
            key="main_results_table",
            on_select="rerun",
            selection_mode="multi-row"
        )

        selected_indices = event.selection.rows
        selected = view_df.iloc[selected_indices] if selected_indices else pd.DataFrame()

        if st.button("➕ Dodaj uređaj"):
            add_selected(selected)

        st.download_button(
            "📥 Preuzmi filtrirani CMDB",
            data=to_excel(view_df),
            file_name="cmdb_pregled.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
else:
    st.info("Unesi parametar za pretragu da bi se prikazali rezultati.")

# =========================
# POSLEDNJA ŠANSA
# =========================
st.markdown("---")
st.subheader("🔍 Poslednja šansa")
st.caption("Ako nisi našao uređaj pokušaj još jednom ovde")

last = st.text_input("Pretraga po svim kolonama", key="last_chance_search")

if st.button("🔎 Pretraži poslednja šansa"):
    if last:
        excluded = [
            "Description",
            "Owner",
            "WarrantyExpirationDate",
            "WarrantzExpirationDate",
            "InstallDate",
            "Note"
        ]

        search_df = df.drop(
            columns=[c for c in excluded if c in df.columns],
            errors="ignore"
        )

        mask = search_df.apply(
            lambda row: row.astype(str).str.lower().str.contains(last.lower()).any(),
            axis=1
        )

        st.session_state.last_chance_results = df[mask].copy()
    else:
        st.session_state.last_chance_results = pd.DataFrame()

if not st.session_state.last_chance_results.empty:
    st.subheader(f"📦 Rezultati poslednje šanse: {len(st.session_state.last_chance_results)}")

    display_cols = [
        "Name", "Vendor", "Model", "Type",
        "SPInventoryNumber", "InventoryNumber", "SerialNumber"
    ]

    available_cols = [
        c for c in display_cols
        if c in st.session_state.last_chance_results.columns
    ]

    view_df = st.session_state.last_chance_results[available_cols].copy()

    event_last = st.dataframe(
        view_df,
        use_container_width=True,
        hide_index=True,
        height=300,
        key="last_chance_results_table",
        on_select="rerun",
        selection_mode="multi-row"
    )

    selected_last_indices = event_last.selection.rows
    selected_last = view_df.iloc[selected_last_indices] if selected_last_indices else pd.DataFrame()

    if st.button("➕ Dodaj iz poslednje šanse"):
        add_selected(selected_last)
else:
    st.info("Klikni na 'Pretraži poslednja šansa' da prikaže rezultate.")

# =========================
# LISTA ZA INTERNI PRENOS
# =========================
st.markdown("---")
st.subheader("🔁 Lista za interni prenos")

if st.session_state.transfer_list:
    header = st.columns([2, 2, 2, 2, 2, 1])
    header[0].markdown("**Name**")
    header[1].markdown("**Model**")
    header[2].markdown("**SP**")
    header[3].markdown("**Inventory**")
    header[4].markdown("**Serial**")
    header[5].markdown("**Ukloni**")

    for i, row in enumerate(st.session_state.transfer_list):
        c = st.columns([2, 2, 2, 2, 2, 1])
        c[0].write(row.get("Name", ""))
        c[1].write(row.get("Model", ""))
        c[2].write(row.get("SPInventoryNumber", ""))
        c[3].write(row.get("InventoryNumber", ""))
        c[4].write(row.get("SerialNumber", ""))

        if c[5].button("🗑️", key=f"del{i}"):
            st.session_state.transfer_list.pop(i)
            st.rerun()

    st.info(f"Ukupno uređaja za prenos: {len(st.session_state.transfer_list)}")
else:
    st.info("Lista je prazna.")

# =========================
# AKCIJE
# =========================
col1, col2, col3 = st.columns(3)

with col1:
    if st.button("BG → NS"):
        generate_internal_transfer(st.session_state.transfer_list, "BG_NS")

with col2:
    if st.button("NIŠ → NS"):
        generate_internal_transfer(st.session_state.transfer_list, "NIS_NS")

with col3:
    if st.button("Obriši listu"):
        st.session_state.transfer_list = []
        st.session_state.generated_excel = None
        st.session_state.generated_file_name = ""
        st.session_state.print_html = ""
        st.rerun()

# =========================
# DOWNLOAD + PRINT
# =========================
if st.session_state.generated_excel:
    st.markdown("---")
    st.subheader("📄 Dokument")

    d1, d2 = st.columns(2)

    with d1:
        st.download_button(
            "📥 Download Excel",
            st.session_state.generated_excel,
            st.session_state.generated_file_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    with d2:
        if st.button("🖨️ Print dokument"):
            components.html(st.session_state.print_html, height=800)